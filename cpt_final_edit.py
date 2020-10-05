from tkinter import *
import math
import sys
from openpyxl import *
from random import *
import random
import time
#import bAJobAM

#subinfo = bAJobAM.infoDialog()

#excel manipulation
name = input("이름 : ")

# create Workbook object
wb = Workbook()
# set file path
filepath = "temp.xlsx"
# save workbook 
ws1 = wb.active
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
ws1.title = 'Raw'
ws2.title = 'Processed'
ws3.title = 'Calculations'

ws1.cell(row = 1, column = 1).value = "Block"
ws1.cell(row = 1, column = 2).value = "Trial"
ws1.cell(row = 1, column = 3).value = "Letter"
ws1.cell(row = 1, column = 4).value = "ISI"
ws1.cell(row = 1, column = 5).value = "Button Hit"
ws1.cell(row = 1, column = 6).value = "X Hit"
ws1.cell(row = 1, column = 7).value = "Reaction Time"

ws2.merge_cells('B1:I1')
ws2.merge_cells('B2:C2')
ws2.merge_cells('D2:E2')
ws2.merge_cells('F2:G2')
ws2.merge_cells('H2:I2')
ws2.merge_cells('J2:K2')
ws2['B1'] = "Time Block"
ws2['B2'] = "1"
ws2['D2'] = "2"
ws2['F2'] = "3"
ws2['H2'] = "4"
ws2['J2'] = "Overall"
ws2['A3'] = "RT"
ws2['A4'] = "RT (SE)"
ws2['A5'] = "Omission Errors (%)"
ws2['A6'] = "Commission Errors (%)"
ws2['A7'] = "d'"
ws2['A8'] = "ß"

ws3.merge_cells('B1:E1')
ws3['B1'] = "Time Block"
ws3['B2'] = "1"
ws3['C2'] = "2"
ws3['D2'] = "3"
ws3['E2'] = "4"
ws3['F2'] = "Overall"
ws3['A3'] = "Total"
ws3['A4'] = "Total X"
ws3['A5'] = "Total Non-X"
ws3['A6'] = "Total Action (Space Button Push)"
ws3['A7'] = "Total Non-action (Pass)"
ws3['A8'] = "Total Omission"
ws3['A9'] = "Total Commission"
ws3['A10'] = "Hit Rate (Probability)"
ws3['A11'] = "Hit Rate (Z-score)"
ws3['A12'] = "Hit Rate (Z-score^2)"
ws3['A13'] = "False Alarm Rate (Probability)"
ws3['A14'] = "False Alarm Rate (Z-Score)"
ws3['A15'] = "False Alarm Rate (Z-score^2)"

####################################
# init
####################################

def init(data):
    # There is only one init, not one-per-mode
    data.mode = "menuScreen"
    data.score = 0
    data.xCount = 0
    data.elseCount = 0
    data.block = 1
    data.timerCalls = 0
    data.lagCalls = 0
    data.lettersDeck = "A B C D E F G H I J K L M N O P Q R S T U V W Y Z".split()
    data.X = "X"
    data.trials = []
    data.trialNum = 0
    data.timerRandom = sample([10, 20, 40], 3)
    data.excelRow = 2
    data.lag = False
    data.initial = 0
    data.final = 0
    data.react_time = 0
    data.allow = True
    data.rt = []
    data.rt_calc = 0
    data.comms = 0

def shuffle(data):
    i = 0
    while (i < 3):
        l = sample(data.lettersDeck, 9)
        l.append("X")
        l = sample(l, 10)
        data.trials.extend(l)
        i += 1

def reset(data):
    data.trials = []
    data.timerCalls = 0
    data.trialNum = 0
    data.timerRandom = sample([10, 20, 40], 3)
    data.score = 0
    data.lagCalls = 0
    data.lag = False
    data.initial = 0
    data.final = 0
    data.react_time = 0
    data.allow = True
    data.comms = 0

####################################
# mode dispatcher
####################################

def mousePressed(event, data):
    if (data.mode == "menuScreen"): menuScreenMousePressed(event, data)
    elif (data.mode == "startTest"):   startTestMousePressed(event, data)
    elif (data.mode == "help"):       helpMousePressed(event, data)
    elif (data.mode == "blockScreen"):  blockScreenMousePressed(event, data)       
    elif (data.mode == "scoreScreen"): scoreScreenMousePressed(event, data)

def keyPressed(event, data):
    if (data.mode == "menuScreen"): menuScreenKeyPressed(event, data)
    elif (data.mode == "startTest"):   startTestKeyPressed(event, data)
    elif (data.mode == "help"):       helpKeyPressed(event, data)
    elif (data.mode == "blockScreen"):       blockScreenKeyPressed(event, data)
    elif (data.mode == "scoreScreen"): scoreScreenKeyPressed(event, data)

def timerFired(data):
    if (data.mode == "menuScreen"): menuScreenTimerFired(data)
    elif (data.mode == "startTest"):   startTestTimerFired(data)
    elif (data.mode == "help"):       helpTimerFired(data)
    elif (data.mode == "blockScreen"):  blockScreenTimerFired(data)
    elif (data.mode == "scoreScreen"): scoreScreenTimerFired(data)

def redrawAll(canvas, data):
    if (data.mode == "menuScreen"): menuScreenRedrawAll(canvas, data)
    elif (data.mode == "startTest"):   startTestRedrawAll(canvas, data)
    elif (data.mode == "help"):       helpRedrawAll(canvas, data)
    elif (data.mode == "blockScreen"):       blockScreenRedrawAll(canvas, data)
    elif (data.mode == "scoreScreen"): scoreScreenRedrawAll(canvas, data)

####################################
# menuScreen mode
####################################

def menuScreenMousePressed(event, data):
    pass

def menuScreenKeyPressed(event, data):
    if (event.keysym == "space"):
        data.mode = "startTest"
        shuffle(data)

def menuScreenTimerFired(data):
    pass

def menuScreenRedrawAll(canvas, data):
    canvas.create_text(data.width/2, data.height/2-20,
                       text="Continuance Performance Test (CPT)", font="Arial 26 bold")
    canvas.create_text(data.width/2, data.height/2+20,
                       text="스페이스 바를 눌러 시작하세요.", font="Arial 20")
    #name = StringVar()
    #entry_box = Canvas()
    #entry_box = Entry(canvas, textvariable = name, width = 25, bg = "lightgreen").place(x = data.width/3, y = 50)
    #name_user = name.get()

####################################
# help mode
####################################

def helpMousePressed(event, data):
    pass

def helpKeyPressed(event, data):
    data.mode = "startTest"

def helpTimerFired(data):
    pass

def helpRedrawAll(canvas, data):
    canvas.create_text(data.width/2, data.height/2-40,
                       text="Help Mode", font="Arial 26 bold")
    canvas.create_text(data.width/2, data.height/2-10,
                       text="이용 방법:", font="Arial 20")
    canvas.create_text(data.width/2, data.height/2+15,
                       text="나타난 글자가 X가 아니면 스페이스 바를 누르세요.", font="Arial 20")
    canvas.create_text(data.width/2, data.height/2+40,
                       text="종료할려면 ESC를 누르세요.", font="Arial 20")
    canvas.create_text(data.width/2, data.height/2+75,
                       text="스페이스 바를 눌러 돌아가세요.", font="Arial 20")

####################################
# excel update code
####################################

def updateExcel(data):
    #if (data.block == 1):
    #block1 - processed
    ws2['B3'] = '=IF(Calculations!B6=0, "Pass", ROUND(AVERAGE(Raw!G2:G31),2))'
    ws2['C3'] = '=IF(B3="Pass", "Pass", ROUND(STDEV(Raw!G2:G31),2))'
    ws2['B4'] = '=IF(C3="Pass", "Pass", ROUND(C3/SQRT(Calculations!B6),2))'
    ws2['B5'] = "=ROUND((Calculations!B8/Calculations!B5)*100,2)"
    ws2['B6'] = "=ROUND((Calculations!B9/Calculations!B4)*100,2)"
    ws2['B7'] = "=ROUND(Calculations!B11-Calculations!B14,2)" 
    ws2['B8'] = "=ROUND(EXP(-1*((Calculations!B12-Calculations!B15)*0.5)),2)"
    #block1 - calc
    ws3['B3'] = "=COUNT(Raw!A2:A31)"
    ws3['B4'] = '=COUNTIF(Raw!C2:C31, "X")'
    ws3['B5'] = "=B3-B4"
    ws3['B6'] = "=SUM(Raw!E2:E31)"
    ws3['B7'] = "=B3-B6"
    ws3['B8'] = "=(B5-B6)+B9"
    ws3['B9'] = "=SUM(Raw!F2:F31)"
    ws3['B10'] = "=IF((B6-B9)/B5=0, 1/(2*B5), IF((B6-B9)/B5=1, 1-(1/(2*B5)), (B6-B9)/B5))" #add hit-rate=0
    ws3['B11'] = "=NORMSINV(B10)"
    ws3['B12'] = "=B11^2"
    ws3['B13'] = "=IF(B9/B4=1, 1-1/(2*B4), IF(B9/B4=0, 1/(2*B4), B9/B4))" #add alarm_rate=1
    ws3['B14'] = "=NORMSINV(B13)"
    ws3['B15'] = "=B14^2"

    #if (data.block == 2):
    #block2 - processed
    ws2['D3'] = '=IF(Calculations!C6=0, "Pass", ROUND(AVERAGE(Raw!G32:G61),2))'
    ws2['E3'] = '=IF(D3="Pass", "Pass", ROUND(STDEV(Raw!G32:G61),2))'
    ws2['D4'] = '=IF(E3="Pass", "Pass", ROUND(E3/SQRT(Calculations!C6),2))'
    ws2['D5'] = "=ROUND((Calculations!C8/Calculations!C5)*100,2)"
    ws2['D6'] = "=ROUND((Calculations!C9/Calculations!C4)*100,2)"
    ws2['D7'] = "=ROUND(Calculations!C11-Calculations!C14,2)"
    ws2['D8'] = "=ROUND(EXP(-1*((Calculations!C12-Calculations!C15)*0.5)),2)"
    #block2 - calc
    ws3['C3'] = "=COUNT(Raw!A32:A61)"
    ws3['C4'] = '=COUNTIF(Raw!C32:C61, "X")'
    ws3['C5'] = "=C3-C4"
    ws3['C6'] = "=SUM(Raw!E32:E61)"
    ws3['C7'] = "=C3-C6"
    ws3['C8'] = "=(C5-C6)+C9"
    ws3['C9'] = "=SUM(Raw!F32:F61)"
    ws3['C10'] = "=IF((C6-C9)/C5=0, 1/(2*C5), IF((C6-C9)/C5=1, 1-(1/(2*C5)), (C6-C9)/C5))" #add hit-rate=0
    ws3['C11'] = "=NORMSINV(C10)"
    ws3['C12'] = "=C11^2"
    ws3['C13'] = "=IF(C9/C4=1, 1-1/(2*C4), IF(C9/C4=0, 1/(2*C4), C9/C4))"
    ws3['C14'] = "=NORMSINV(C13)"
    ws3['C15'] = "=C14^2"

    #if (data.block == 3):
    #block3 - processed
    ws2['F3'] = '=IF(Calculations!D6=0, "Pass", ROUND(AVERAGE(Raw!G62:G91),2))'
    ws2['G3'] = '=IF(F3="Pass", "Pass", ROUND(STDEV(Raw!G62:G91),2))'
    ws2['F4'] = '=IF(G3="Pass", "Pass", ROUND(G3/SQRT(Calculations!D6),2))'
    ws2['F5'] = "=ROUND((Calculations!D8/Calculations!D5)*100,2)"
    ws2['F6'] = "=ROUND((Calculations!D9/Calculations!D4)*100,2)"
    ws2['F7'] = "=ROUND(Calculations!D11-Calculations!D14,2)"
    ws2['F8'] = "=ROUND(EXP(-1*((Calculations!D12-Calculations!D15)*0.5)),2)"
    #block3 - calc
    ws3['D3'] = "=COUNT(Raw!A62:A91)"
    ws3['D4'] = '=COUNTIF(Raw!C62:C91, "X")'
    ws3['D5'] = "=D3-D4"
    ws3['D6'] = "=SUM(Raw!E62:E91)"
    ws3['D7'] = "=D3-D6"
    ws3['D8'] = "=(D5-D6)+D9"
    ws3['D9'] = "=SUM(Raw!F62:F91)"
    ws3['D10'] = "=IF((D6-D9)/D5=0, 1/(2*D5), IF((D6-D9)/D5=1, 1-(1/(2*D5)), (D6-D9)/D5))" #add hit-rate=0
    ws3['D11'] = "=NORMSINV(D10)"
    ws3['D12'] = "=D11^2"
    ws3['D13'] = "=IF(D9/D4=1, 1-1/(2*D4), IF(D9/D4=0, 1/(2*D4), D9/D4))"
    ws3['D14'] = "=NORMSINV(D13)"
    ws3['D15'] = "=D14^2"

    #if (data.block == 4):
    #block4 - processed
    ws2['H3'] = '=IF(Calculations!E6=0, "Pass", ROUND(AVERAGE(Raw!G92:G121),2))'
    ws2['I3'] = '=IF(H3="Pass", "Pass", ROUND(STDEV(Raw!G92:G121),2))'
    ws2['H4'] = '=IF(I3="Pass", "Pass", ROUND(I3/SQRT(Calculations!E6),2))'
    ws2['H5'] = "=ROUND((Calculations!E8/Calculations!E5)*100,2)"
    ws2['H6'] = "=ROUND((Calculations!E9/Calculations!E4)*100,2)"
    ws2['H7'] = "=ROUND(Calculations!E11-Calculations!E14,2)"
    ws2['H8'] = "=ROUND(EXP(-1*((Calculations!E12-Calculations!E15)*0.5)),2)"
    #block4 - calc
    ws3['E3'] = "=COUNT(Raw!A92:A121)"
    ws3['E4'] = '=COUNTIF(Raw!C92:C121, "X")'
    ws3['E5'] = "=E3-E4"
    ws3['E6'] = "=SUM(Raw!E92:E121)"
    ws3['E7'] = "=E3-D6"
    ws3['E8'] = "=(E5-E6)+E9"
    ws3['E9'] = "=SUM(Raw!F92:F121)"
    ws3['E10'] = "=IF((E6-E9)/E5=0, 1/(2*E5), IF((E6-E9)/E5=1, 1-(1/(2*E5)), (E6-E9)/E5))" #add hit-rate=0
    ws3['E11'] = "=NORMSINV(E10)"
    ws3['E12'] = "=E11^2"
    ws3['E13'] = "=IF(E9/E4=1, 1-1/(2*E4), IF(E9/E4=0, 1/(2*E4), E9/E4))"
    ws3['E14'] = "=NORMSINV(E13)"
    ws3['E15'] = "=E14^2"

    #overall - processed
    ws2['J3'] = '=IF(Calculations!F6=0, "Pass", ROUND(AVERAGE(Raw!G2:G121),2))'
    ws2['K3'] = '=IF(J3="Pass", "Pass", ROUND(STDEV(Raw!G2:G121),2))'
    ws2['J4'] = '=IF(K3="Pass", "Pass", ROUND(O3/SQRT(Calculations!F6),2))'
    ws2['J5'] = "=ROUND((Calculations!F8/Calculations!F5)*100,2)"
    ws2['J6'] = "=ROUND((Calculations!F9/Calculations!F4)*100,2)"
    ws2['J7'] = "=ROUND(Calculations!F11-Calculations!F14,2)"
    ws2['J8'] = "=ROUND(EXP(-1*((Calculations!F12-Calculations!F15)*0.5)),2)"
    #overall - calc
    ws3['F3'] = "=COUNT(Raw!A2:A121)"
    ws3['F4'] = '=COUNTIF(Raw!C2:C121, "X")'
    ws3['F5'] = "=F3-F4"
    ws3['F6'] = "=SUM(Raw!E2:E121)"
    ws3['F7'] = "=F3-F6"
    ws3['F8'] = "=(F5-F6)+F9"
    ws3['F9'] = "=SUM(Raw!F2:F121)"
    ws3['F10'] = "=IF((F6-F9)/F5=0, 1/(2*F5), IF((F6-F9)/F5=1, 1-(1/(2*F5)), (F6-F9)/F5))" #add hit-rate=0
    ws3['F11'] = "=NORMSINV(F10)"
    ws3['F12'] = "=F11^2"
    ws3['F13'] = "=IF(F9/F4=1, 1-1/(2*F4), IF(F9/F4=0, 1/(2*F4), F9/F4))"
    ws3['F14'] = "=NORMSINV(F13)"
    ws3['F15'] = "=F14^2"

####################################
# startTest mode
####################################

def notX(letter):
    if (letter == "X"): return False
    else: return True

def showLetter(canvas, data):
    data.initial = time.time()
    cur = data.trials[data.trialNum]
    canvas.create_text(data.width/2, data.height/2,
                       text=cur, font="Arial 200 bold")

    #updateExcel(data)

def showCross(canvas, data):
    canvas.create_text(data.width/2, data.height/2, text="+", font="Arial 200 bold")
    updateExcel(data)

def chooseTime(data):
    if (data.trialNum < 10):
        timeLim = data.timerRandom[0]
    elif (10 <= data.trialNum < 20):
        timeLim = data.timerRandom[1]
    elif (20 <= data.trialNum):
        timeLim = data.timerRandom[2]
    return timeLim

def startTestMousePressed(event, data):
    pass

def startTestKeyPressed(event, data):
    excelTrialEq = 30*(data.block-1) + data.trialNum
    if (event.keysym == 'h'):
        data.mode = "help" 
    if (event.keysym == "space" and data.allow):
        data.allow = not data.allow
        data.final = time.time()
        data.react_time = data.final - data.initial
        if (data.trials[data.trialNum] != "X"): 
            data.score += 1
            ws1.cell(row = excelTrialEq + 2, column = 7).value = float("%.2f" % (1000*data.react_time))
        elif (data.trials[data.trialNum] == "X"):
            data.comms += 1
            ws1.cell(row = excelTrialEq + 2, column = 6).value = 1
        ws1.cell(row = excelTrialEq + 2, column = 2).value = excelTrialEq + 1
        ws1.cell(row = excelTrialEq + 2, column = 1).value = data.block
        ws1.cell(row = excelTrialEq + 2, column = 3).value = data.trials[data.trialNum]
        ws1.cell(row = excelTrialEq + 2, column = 5).value = 1

def startTestTimerFired(data):
    timeLim = chooseTime(data)
    excelTrialEq = 30*(data.block-1) + data.trialNum
    ws1.cell(row = excelTrialEq + 2, column = 4).value = float("%.2f" % (timeLim/10)) 
    if (not data.lag):
        data.timerCalls += 1
    elif (data.lag): 
        data.lagCalls += 1
        if (data.lagCalls % timeLim == 0):
            data.lag = not data.lag
            data.lagCalls = 0
            data.trialNum += 1
            print(time.time() - data.initial)
    if (data.timerCalls >= 2):
        if (data.trialNum < 30): data.allow = True
        ws1.cell(row = excelTrialEq + 2, column = 2).value = excelTrialEq + 1
        ws1.cell(row = excelTrialEq + 2, column = 1).value = data.block
        ws1.cell(row = excelTrialEq + 2, column = 3).value = data.trials[data.trialNum]
        ws1.cell(row = excelTrialEq + 2, column = 5).value = 0
        if (data.trials[data.trialNum] == "X"): 
            data.score += 1
        data.lag = not data.lag
        data.timerCalls = 0
    
def startTestRedrawAll(canvas, data):
    if (data.trialNum % 30 == 0 and data.trialNum != 0): 
        data.allow = False
        data.timerCalls = 0
        data.trialNum = 0
        data.block += 1
        reset(data)
        shuffle(data)
    if (data.block == 5):
        data.allow = False
        data.mode = "scoreScreen"
    if (not data.lag):
        showLetter(canvas, data)
    elif (data.lag): 
        showCross(canvas, data)

####################################
# blockScreen mode
####################################

def blockScreenMousePressed(event, data):
    pass

def blockScreenKeyPressed(event, data):
    pass

def blockScreenTimerFired(data):
    data.timerCalls += 1

def blockScreenRedrawAll(canvas, data):
    if (data.timerCalls <= 30):
        canvas.create_text(data.width/2, data.height/2, 
            text="Block" + " " + str(data.block), font="Arial 40")
    if (data.timerCalls == 40):
        reset(data)
        shuffle(data)
        data.mode = "startTest"

####################################
# scoreScreen mode
####################################

def scoreScreenMousePressed(event, data):
    pass

def scoreScreenKeyPressed(event, data):
    pass

def scoreScreenTimerFired(data):
    pass

def scoreScreenRedrawAll(canvas, data):
    canvas.create_text(data.width/2, data.height/2,
                       text="끝", font="Arial 200")
    

####################################
# use the run function as-is
####################################

def run(width=600, height=600):
    def redrawAllWrapper(canvas, data):
        canvas.delete(ALL)
        canvas.create_rectangle(0, 0, data.width, data.height,
                                fill='#f2f2f2', width=0)
        redrawAll(canvas, data)
        canvas.update()

    def mousePressedWrapper(event, canvas, data):
        mousePressed(event, data)
        redrawAllWrapper(canvas, data)

    def keyPressedWrapper(event, canvas, data):
        keyPressed(event, data)
        redrawAllWrapper(canvas, data)

    def timerFiredWrapper(canvas, data):
        timerFired(data)
        redrawAllWrapper(canvas, data)
        # pause, then call timerFired again
        canvas.after(data.timerDelay, timerFiredWrapper, canvas, data)

    # Set up data and call init
    class Struct(object): pass
    data = Struct()
    data.width = width
    data.height = height
    data.timerDelay = 89 #milliseconds
    root = Tk()
    init(data)
    # create the root and the canvas
    canvas = Canvas(root, width=data.width, height=data.height)
    canvas.pack()
    # set up events
    root.bind("<Button-1>", lambda event:
                            mousePressedWrapper(event, canvas, data))
    root.bind("<Key>", lambda event:
                            keyPressedWrapper(event, canvas, data))
    timerFiredWrapper(canvas, data)
    #close
    def close(event):
        root.destroy()
    root.bind('<Escape>',close)
    # and launch the app
    root.mainloop()  # blocks until window is closed
    print("bye!")

run(600,600)

wb.save(name + ".xlsx")
#wb.save(subinfo[1]+subinfo[0] + ".xlsx")